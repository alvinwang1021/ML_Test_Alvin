from metric_learn import LMNN
from metric_learn import Covariance
from metric_learn import ITML_Supervised
from metric_learn import LFDA
from metric_learn import NCA
from metric_learn import SDML_Supervised
from metric_learn import RCA_Supervised
import numpy as np
#import scipy
#from sklearn.preprocessing import scale
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split
import datetime

LColumn = 4

#Method = 'LMNN'
#Method = 'ITML'
#Method = 'COV'
#Method = 'LFDA'
#Method = 'NCA'
Method = 'SDML'
#Method = 'RCA'
#

wb = load_workbook(filename = 'demographic+Data+From+mimic.xlsx')
ws = wb.active
colNo = len(list(ws.columns))
rowNo = len(list(ws.rows))

i = 0
for row in ws.iter_rows(min_row=2, max_col = colNo, max_row=rowNo):
    i = i + 1
    curRow =list()
    for cell in row:
        if cell.value is None:
            cell.value = "Missing Value"
        curRow.append(cell.value)
    if i == 1:
        Data = np.array([curRow],dtype=object)
    else:
        Data = np.insert(Data, len(Data), values=curRow, axis=0)



#print(len(TrainData.T[15]),'\n')
for i in range (0, len(Data)):
    if isinstance(Data.T[8][i], datetime.date):
        Data.T[8][i] = '0 Day'
    for s in (Data.T[8][i]).split():
        if s.isdigit():
            Data.T[8][i] = int(s)


from sklearn import preprocessing
le = preprocessing.LabelEncoder()
LabelName = list()
LabelCode = list()
for i in range (0,8):
    le.fit(Data.T[i])
    #print(list(le.classes_))
    LabelName.append(list(le.classes_))
    LabelCode.append(list(le.transform(list(le.classes_))))
    

    Data.T[i] = le.transform(Data.T[i]) 
    #print(TrainData[2])
    #array([2, 2, 1]...)
    #print(list(le.inverse_transform([2, 2, 1])))
#print(TrainData)
#print('sssss2', len(TrainData[0]))
'''
print (LabelName[7])
print (LabelCode[7])
print (len(LabelName[7]))
print (len(LabelCode[7]))
'''


#print(TrainData[0:5])

def column(matrix, i):
    return [row[i] for row in matrix]
    
labels = np.array(column(Data, LColumn))
Data = np.delete(Data, (LColumn), axis = 1)
#print(labels,'\n')
#print(len(labels))
if LColumn < 8:
    print(LabelName[LColumn],'\n')

#print('sssss3', len(TrainData[0]))
#print(TrainData)

#print(len(Trainlabels))
#print(len(TrainData[0]))
#print(TrainData[0:8])
'''
OneHotEncoder
'''

from sklearn.preprocessing import OneHotEncoder
if LColumn < 8:
    enc = OneHotEncoder(categorical_features= np.array([0,1,2,3,4,5,6]))
    enc.fit(Data)
    FinalData = enc.transform(Data).toarray()
else:
    enc = OneHotEncoder(categorical_features= np.array([0,1,2,3,4,5,6,7]))
    enc.fit(Data)
    FinalData = enc.transform(Data).toarray()


TrainData, TestData, TrainLabels, TestLabels = train_test_split(FinalData, labels, test_size=0.3)

#STrainData = scale(TrainData)
#STestData = scale(TestData)

min_max_scaler = MinMaxScaler()
STrainData = min_max_scaler.fit_transform(TrainData)

min_max_scaler = MinMaxScaler()
STestData = min_max_scaler.fit_transform(TestData)
#print(STrainData.max(axis=0) - STrainData.min(axis=0))
#print (STrainData.mean(axis = 0))
#print (STrainData.std(axis = 0))

#print('max',STrainData.max(axis=0))
#print('min',STrainData.min(axis=0))
#print(STrainData.max(axis=0) - STrainData.min(axis=0))
#print(len(STrainData.max(axis=0) - STrainData.min(axis=0)))
#print(TestData.max(axis=0) - TestData.min(axis=0))
#print(STrainData.max(axis=0) - STrainData.min(axis=0))
count = 0
for i in range (len(STrainData.max(axis=0) - STrainData.min(axis=0))):
    if (STrainData.max(axis=0) - STrainData.min(axis=0))[i] == 0:
        if count == 0:
            col = [i]
            count = count + 1
        else:
            col.append(i)

#print(col)
try: 
    FSTrainData = np.delete(STrainData, col, axis = 1)
    FSTestData = np.delete(STestData, col, axis = 1)
except:
    FSTrainData = STrainData
    FSTestData = STestData

print('Data Preparation Done', '\n')
#print(FSTrainData.max(axis=0) - FSTrainData.min(axis=0))

#print(len(FSTrainData[0]))
#print(len(FSTestData[0]))
#print(len(FSTestData))
#print(len(TestData))
#print(TrainData)
#print(type(TrainData))
#print(TrainLabels)
#print(type(TrainLabels))


if Method == 'LMNN':
    print("Method: LMNN", '\n')
    lmnn = LMNN(k=3, learn_rate=1e-6, verbose=False)
    x = lmnn.fit(FSTrainData, TrainLabels)
    TFSTestData = x.transform(FSTestData)
    print('Transformation Done', '\n')

elif Method == 'COV':
    print("Method: COV", '\n')
    cov = Covariance().fit(FSTrainData)
    TFSTestData = cov.transform(FSTestData)
    print('Transformation Done', '\n')

elif Method == 'ITML':
    print("Method: ITML", '\n')
    itml = ITML_Supervised(num_constraints=200,A0=None)
    x = itml.fit(FSTrainData, TrainLabels)
    TFSTestData = x.transform(FSTestData)
    print('Transformation Done', '\n')

elif Method == 'LFDA':
    print("Method: LFDA", '\n')
    lfda = LFDA(k=4, dim=1)
    x = lfda.fit(FSTrainData, TrainLabels)
    TFSTestData = x.transform(FSTestData)
    print('Transformation Done', '\n')

elif Method == 'NCA':
    print("Method: NCA", '\n')
    #print('Max', TrainData.max(axis=0))
    #print('sssssssss', len(TrainData[0]))
    #print('sssssssss', len(TrainData.max(axis=0)))
    #print('Min', TrainData.min(axis=0))

    nca = NCA(max_iter=500, learning_rate=0.01)
    # print('ssssssss', TrainData)
    x = nca.fit(FSTrainData, TrainLabels)
    
    TFSTestData = x.transform(FSTestData)
    print('Transformation Done', '\n')

elif Method == 'SDML':
    print("Method: SDML", '\n')
    sdml = SDML_Supervised(num_constraints=200)
    x= sdml.fit(FSTrainData, TrainLabels)
    TFSTestData = x.transform(FSTestData)
    print('Transformation Done', '\n')

elif Method == 'RCA':
    print("Method: RCA", '\n')
    rca = RCA_Supervised(num_chunks=2, chunk_size=1)
    x= rca.fit(FSTrainData, TrainLabels)
    TFSTestData = x.transform(FSTestData)
    print('Transformation Done', '\n')

#print(len(TTestData))
#print(TTestData[0])

#rca = RCA_Supervised(num_chunks=2, chunk_size=1)
#x= rca.fit(TrainData, targets)

#TTestData = x.transform(TestData)
#transformer = x.transformer()
#print(TTestData)

def EuclDist(coords1, coords2):

   dist = 0
   for (x, y) in zip(coords1, coords2):
        dist += (x - y)**2
   return dist**0.5

#print(LColumn)
#print('ssssssss', TestLabels[98])
#print(TestLabels[0:100])
#print(len(TestLabels))
for i in range (0, len(TestLabels)):
    #print(i)
    if TestLabels[i] == 0:
        ClassA = list()
        TClassA = list()
    elif TestLabels[i] == 1:
        ClassB = list()
        TClassB = list()
    elif TestLabels[i] == 2:
        ClassC = list()
        TClassC = list()
    elif TestLabels[i] == 3:
        ClassD = list()
        TClassD = list()
    elif TestLabels[i] == 4:
        ClassE = list()
        TClassE = list()
    elif TestLabels[i] == 5:
        ClassF = list()
        TClassF = list()
    elif TestLabels[i] == 6:
        ClassG = list()
        TClassG = list()
    elif TestLabels[i] == 7:
        ClassH = list()
        TClassH = list()
    elif TestLabels[i] == 8:
        ClassI = list()
        TClassI = list()
    elif TestLabels[i] == 9:
        ClassJ = list()
        TClassJ = list()

for i in range (0, len(TestLabels)):
    if TestLabels[i] == 0:
        #print(TestData[i], '\n')
        ClassA.append(FSTestData[i])
        TClassA.append(TFSTestData[i])
    elif TestLabels[i] == 1:
        ClassB.append(FSTestData[i])
        TClassB.append(TFSTestData[i])
    elif TestLabels[i] == 2:
        ClassC.append(FSTestData[i])
        TClassC.append(TFSTestData[i])
    elif TestLabels[i] == 3:
        ClassD.append(FSTestData[i])
        TClassD.append(TFSTestData[i])
    elif TestLabels[i] == 4:
        ClassE.append(FSTestData[i])
        TClassE.append(TFSTestData[i])
    elif TestLabels[i] == 5:
        ClassF.append(FSTestData[i])
        TClassF.append(TFSTestData[i])
    elif TestLabels[i] == 6:
        ClassG.append(FSTestData[i])
        TClassG.append(TFSTestData[i])
    elif TestLabels[i] == 7:
        ClassH.append(FSTestData[i])
        TClassH.append(TFSTestData[i])
    elif TestLabels[i] == 8:
        ClassI.append(FSTestData[i])
        TClassI.append(TFSTestData[i])
    elif TestLabels[i] == 9:
        ClassJ.append(FSTestData[i])
        TClassJ.append(TFSTestData[i])

try: 
    ClassA = np.array(ClassA)
    ClassACenter = sum(ClassA)/len(ClassA)
    TClassA = np.array(TClassA)
    TClassACenter = sum(TClassA)/len(TClassA)
except: pass

try:
    ClassB = np.array(ClassB)
    ClassBCenter = sum(ClassB)/len(ClassB)
    TClassB = np.array(TClassB)
    TClassBCenter = sum(TClassB)/len(TClassB)
except: pass

try: 
    ClassC = np.array(ClassC)
    ClassCCenter = sum(ClassC)/len(ClassC)
    TClassC = np.array(TClassC)
    TClassCCenter = sum(TClassC)/len(TClassC)
except: pass
try: 
    ClassD = np.array(ClassD)
    ClassDCenter = sum(ClassD)/len(ClassD)
    TClassD = np.array(TClassD)
    TClassDCenter = sum(TClassD)/len(TClassD)
except: pass
try: 
    ClassE = np.array(ClassE)
    ClassECenter = sum(ClassE)/len(ClassE)
    TClassE = np.array(TClassE)
    TClassECenter = sum(TClassE)/len(TClassE)
except: pass
try: 
    ClassF = np.array(ClassF)
    ClassFCenter = sum(ClassF)/len(ClassF)
    TClassF = np.array(TClassF)
    TClassFCenter = sum(TClassF)/len(TClassF)
except: pass    
try: 
    ClassG = np.array(ClassG)
    ClassGCenter = sum(ClassG)/len(ClassG)
    TClassG = np.array(TClassG)
    TClassGCenter = sum(TClassG)/len(TClassG)
except: pass    
try: 
    ClassH = np.array(ClassH)
    ClassHCenter = sum(ClassH)/len(ClassH)
    TClassH = np.array(TClassH)
    TClassHCenter = sum(TClassH)/len(TClassH)
except: pass    
try: 
    ClassI = np.array(ClassI)
    ClassICenter = sum(ClassI)/len(ClassI)
    TClassI = np.array(TClassI)
    TClassICenter = sum(TClassI)/len(TClassI)
except: pass
try: 
    ClassJ = np.array(ClassJ)
    ClassJCenter = sum(ClassJ)/len(ClassJ)
    TClassJ = np.array(TClassJ)
     
except: pass


OCorrect = 0
TCorrect = 0
#for i in random.sample(range(149), 149):

for i in range(0, len(FSTestData)):
    Odistance = list()
    Tdistance = list()
    
    try:
        OdistanceA = EuclDist(FSTestData[i], ClassACenter)
        Odistance.append(OdistanceA)
        TdistanceA = EuclDist(TFSTestData[i], TClassACenter)
        Tdistance.append(TdistanceA)
    except: pass
    
    try:
        OdistanceB = EuclDist(FSTestData[i], ClassBCenter)
        Odistance.append(OdistanceB)
        TdistanceB = EuclDist(TFSTestData[i], TClassBCenter)
        Tdistance.append(TdistanceB)
    except: pass
    
    try:
        OdistanceC = EuclDist(FSTestData[i], ClassCCenter)
        Odistance.append(OdistanceC)
        TdistanceC = EuclDist(TFSTestData[i], TClassCCenter)
        Tdistance.append(TdistanceC)

    except: pass
    try:
        OdistanceD = EuclDist(FSTestData[i], ClassDCenter)
        Odistance.append(OdistanceD)
        TdistanceD = EuclDist(TFSTestData[i], TClassDCenter)
        Tdistance.append(TdistanceD)
    except: pass    
    try:
        OdistanceE = EuclDist(FSTestData[i], ClassECenter)
        Odistance.append(OdistanceE)
        TdistanceE = EuclDist(TFSTestData[i], TClassECenter)
        Tdistance.append(TdistanceE)
    except: pass
    try:
        OdistanceF = EuclDist(FSTestData[i], ClassFCenter)
        Odistance.append(OdistanceF)
        TdistanceF = EuclDist(TFSTestData[i], TClassFCenter)
        Tdistance.append(TdistanceF)
    except: pass
    try:
        OdistanceG = EuclDist(FSTestData[i], ClassGCenter)
        Odistance.append(OdistanceG)
        TdistanceG = EuclDist(TFSTestData[i], TClassGCenter)
        Tdistance.append(TdistanceG)
    except: pass
    try:
        OdistanceH = EuclDist(FSTestData[i], ClassHCenter)
        Odistance.append(OdistanceH)
        TdistanceH = EuclDist(TFSTestData[i], TClassHCenter)
        Tdistance.append(TdistanceH)
    except: pass
    try:
        OdistanceI = EuclDist(FSTestData[i], ClassICenter)
        Odistance.append(OdistanceI)
        TdistanceI = EuclDist(TFSTestData[i], TClassICenter)
        Tdistance.append(TdistanceI)
    except: pass
    try:
        OdistanceJ = EuclDist(FSTestData[i], ClassJCenter)
        Odistance.append(OdistanceJ)
        TdistanceJ = EuclDist(TFSTestData[i], TClassJCenter)
        Tdistance.append(TdistanceJ)
    except: pass


    #print(min(OdistanceA,OdistanceB,OdistanceC))
    try:
        if min(Odistance) == OdistanceA:
            OClassfication = 0
        if min(Tdistance) == TdistanceA:
            TClassfication = 0
    except: pass
    
    try:
        if min(Odistance) == OdistanceB:
            OClassfication = 1
        if min(Tdistance) == TdistanceB:
            TClassfication = 1
    except: pass
    
    try:
        if min(Odistance) == OdistanceC:
            OClassfication = 2
        if min(Tdistance) == TdistanceC:
            TClassfication = 2
    except: pass
    try:
        if min(Odistance) == OdistanceD:
            OClassfication = 3
        if min(Tdistance) == TdistanceD:
            TClassfication = 3
    except: pass
    try:
        if min(Odistance) == OdistanceE:
            OClassfication = 4
        if min(Tdistance) == TdistanceE:
            TClassfication = 4
    except: pass
    try:
        if min(Odistance) == OdistanceF:
            OClassfication = 5
        if min(Tdistance) == TdistanceF:
            TClassfication = 5
    except: pass
    try:
        if min(Odistance) == OdistanceG:
            OClassfication = 6
        if min(Tdistance) == TdistanceG:
            TClassfication = 6
    except: pass
    try:
        if min(Odistance) == OdistanceH:
            OClassfication = 7
        if min(Tdistance) == TdistanceH:
            TClassfication = 7
    except: pass
    try:
        if min(Odistance) == OdistanceI:
            OClassfication = 8
        if min(Tdistance) == TdistanceI:
            TClassfication = 8
    except: pass
    try:
        if min(Odistance) == OdistanceJ:
            OClassfication = 9
        if min(Tdistance) == TdistanceJ:
            TClassfication = 9
    except: pass

    if OClassfication == TestLabels[i]:
        OCorrect = OCorrect+1
    if TClassfication == TestLabels[i]:
        TCorrect = TCorrect+1

#print('TClassB: ', TClassB)
#print('TClassBCenter: ', TClassBCenter, '\n', '\n')

print('Total Samples for Training: ',len(TrainData), '\n')
print('Total Samples for Test: ',len(TestData),'\n')

print('Correct Test Samples in original: ',OCorrect)
#print(len(TestData))
Oaccuracy = OCorrect / len(TestData)
print('Accuracy in original: ',100*Oaccuracy, '%','\n') 

print('Correct Test Samples after transformation: ',TCorrect)
Taccuracy = TCorrect / len(TestData)
print('Accuracy after transformation: ', 100*Taccuracy, '%', '\n') 
print('Accuracy increased by ', (Taccuracy-Oaccuracy)*100, '%', '\n')
